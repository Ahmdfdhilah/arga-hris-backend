"""
Excel parser utility untuk bulk import
Mendukung parsing Excel file untuk org units dan employees
"""
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO


class ExcelParser:
    """Utility class untuk parsing Excel files"""

    @staticmethod
    def parse_org_units_sheet(file_content: bytes, sheet_name: str = "Department") -> List[Dict[str, Any]]:
        """
        Parse Excel sheet untuk org units

        Mapping kolom Excel 'Department':
        - Kode -> code
        - Nama -> name
        - Tipe -> type
        - Head Departement -> parent_code
        - Head Email -> head_email
        - Deskripsi -> description

        Args:
            file_content: Binary content dari Excel file
            sheet_name: Nama sheet (default: "Department")

        Returns:
            List of dictionaries dengan mapped fields

        Raises:
            ValueError: Jika sheet tidak ditemukan atau format tidak valid
        """
        try:
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {str(e)}")

        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")

        sheet: Worksheet = workbook[sheet_name]

        # Get header row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        # Mapping kolom Indonesia ke English
        column_mapping = {
            "Kode": "code",
            "Nama": "name",
            "Tipe": "type",
            "Head Department": "parent_code",
            "Head Email": "head_email",
            "Deskripsi": "description"
        }

        # Validate required columns
        required_columns = ["Kode", "Nama", "Tipe"]
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

        # Parse rows
        results = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_data = {}
            for idx, header in enumerate(headers):
                if header in column_mapping:
                    value = row[idx] if idx < len(row) else None
                    # Convert to string and strip whitespace
                    if value is not None:
                        value = str(value).strip()
                        # Convert empty strings to None
                        if value == "" or value.lower() == "none":
                            value = None
                    row_data[column_mapping[header]] = value

            # Add row number for error tracking
            row_data["row_number"] = row_idx

            # Skip if required fields are missing
            if not row_data.get("code") or not row_data.get("name") or not row_data.get("type"):
                continue

            results.append(row_data)

        return results

    @staticmethod
    def parse_employees_sheet(file_content: bytes, sheet_name: str = "Karyawan") -> List[Dict[str, Any]]:
        """
        Parse Excel sheet untuk employees

        Mapping kolom Excel 'Karyawan':
        - Nomor -> number
        - Nama Depan -> first_name
        - Nama Belakang -> last_name
        - Email -> email
        - Department -> org_unit_name
        - Nomor HP -> phone
        - Jabatan -> position
        - Tipe Karyawan -> employee_type
        - Jenis Karyawan -> account_type
        - Gender -> employee_gender
        - Awal Kontrak -> valid_from
        - Selesai Kontrak -> valid_until
        - Catatan -> notes

        Args:
            file_content: Binary content dari Excel file
            sheet_name: Nama sheet (default: "Karyawan")

        Returns:
            List of dictionaries dengan mapped fields

        Raises:
            ValueError: Jika sheet tidak ditemukan atau format tidak valid
        """
        try:
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {str(e)}")

        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")

        sheet: Worksheet = workbook[sheet_name]

        # Get header row
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        # Mapping kolom Indonesia ke English
        column_mapping = {
            "Nomor": "number",
            "Nama Depan": "first_name",
            "Nama Belakang": "last_name",
            "Email": "email",
            "Department": "org_unit_name",
            "Nomor HP": "phone",
            "Jabatan": "position",
            "Tipe Akun": "account_type",
            "Jenis Karyawan": "employee_type",
            "Gender": "employee_gender",
            "Awal Kontrak": "valid_from",
            "Selesai Kontrak": "valid_until",
            "Catatan": "notes",
        }

        # Validate required columns
        required_columns = ["Nomor", "Nama Depan", "Nama Belakang", "Email", "Department"]
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

        # Parse rows
        results = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_data = {}
            for idx, header in enumerate(headers):
                if header in column_mapping:
                    value = row[idx] if idx < len(row) else None
                    # Convert to string and strip whitespace
                    if value is not None:
                        value = str(value).strip()
                        # Convert empty strings to None
                        if value == "" or value.lower() == "none":
                            value = None
                    row_data[column_mapping[header]] = value

            # Add row number for error tracking
            row_data["row_number"] = row_idx


            # Skip if required fields are missing
            if (not row_data.get("number") or
                not row_data.get("first_name") or
                not row_data.get("last_name") or
                not row_data.get("email") or
                not row_data.get("org_unit_name")):
                continue

            # Set default account_type if not provided
            if not row_data.get("account_type"):
                row_data["account_type"] = "user"

            results.append(row_data)

        return results

    @staticmethod
    def validate_org_units_data(data: List[Dict[str, Any]]) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Validate org units data and separate valid/invalid items

        Args:
            data: List of org unit dictionaries

        Returns:
            Tuple of (valid_items, invalid_items)
        """
        valid_items = []
        invalid_items = []

        for item in data:
            errors = []

            # Validate code
            if not item.get("code"):
                errors.append("Kode tidak boleh kosong")

            # Validate name
            if not item.get("name"):
                errors.append("Nama tidak boleh kosong")

            # Validate type
            if not item.get("type"):
                errors.append("Tipe tidak boleh kosong")

            if errors:
                invalid_items.append({
                    "row_number": item.get("row_number"),
                    "code": item.get("code"),
                    "errors": errors
                })
            else:
                valid_items.append(item)

        return valid_items, invalid_items

    @staticmethod
    def validate_employees_data(data: List[Dict[str, Any]]) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Validate employees data and separate valid/invalid items

        Args:
            data: List of employee dictionaries

        Returns:
            Tuple of (valid_items, invalid_items)
        """
        valid_items = []
        invalid_items = []

        for item in data:
            errors = []

            # Validate required fields
            if not item.get("number"):
                errors.append("Nomor karyawan tidak boleh kosong")

            if not item.get("first_name"):
                errors.append("Nama depan tidak boleh kosong")

            if not item.get("last_name"):
                errors.append("Nama belakang tidak boleh kosong")

            if not item.get("email"):
                errors.append("Email tidak boleh kosong")

            if not item.get("org_unit_name"):
                errors.append("Kode unit organisasi tidak boleh kosong")

            # Validate email format (basic)
            if item.get("email") and "@" not in item.get("email", ""):
                errors.append("Format email tidak valid")

            # Validate employee_type if provided
            if item.get("employee_type") and item.get("employee_type") not in ["on_site", "hybrid"]:
                errors.append("Tipe karyawan harus 'on_site' atau 'hybrid'")

            # Validate employee_gender if provided
            if item.get("employee_gender") and item.get("employee_gender") not in ["male", "female"]:
                errors.append("Gender harus 'male' atau 'female'")

            # Validate account_type
            if item.get("account_type") and item.get("account_type") not in ["user", "guest", "none"]:
                errors.append("Jenis karyawan harus 'user', 'guest', atau 'none'")

            if errors:
                invalid_items.append({
                    "row_number": item.get("row_number"),
                    "number": item.get("number"),
                    "errors": errors
                })
            else:
                valid_items.append(item)

        return valid_items, invalid_items
